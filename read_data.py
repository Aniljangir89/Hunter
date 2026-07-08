import openpyxl
import os

# Find the file
data_dir = "Data"
files = os.listdir(data_dir)
print("Files in Data dir:", files)
filepath = os.path.join(data_dir, files[0])
print("Loading:", filepath)

wb = openpyxl.load_workbook(filepath)
print("Sheets:", wb.sheetnames)
ws = wb.active
print("Active sheet:", ws.title)
print("Rows:", ws.max_row, "Cols:", ws.max_column)
print("\n--- Headers ---")
for cell in ws[1]:
    print(f"  Col {cell.column}: {cell.value}")
print("\n--- Sample rows (2-6) ---")
for row in ws.iter_rows(min_row=2, max_row=6):
    print([cell.value for cell in row])

# Check all sheets
for name in wb.sheetnames:
    sheet = wb[name]
    print(f"\n=== Sheet: {name} === Rows: {sheet.max_row}, Cols: {sheet.max_column}")
    for cell in sheet[1]:
        print(f"  Col {cell.column}: {cell.value}")
    print("Sample rows:")
    for row in sheet.iter_rows(min_row=2, max_row=4):
        print([cell.value for cell in row])
